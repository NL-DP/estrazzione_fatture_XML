from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from extractor_fatture.config import SHEET_RIEPILOGO


def read_processed_invoice_ids(output_path: Path) -> set[str]:
    """
    Legge il foglio Riepilogo dell'Excel esistente e restituisce
    l'insieme degli invoice_id già presenti.
    """
    if not output_path.exists():
        return set()

    try:
        wb = load_workbook(str(output_path), read_only=True, data_only=True)

        if SHEET_RIEPILOGO not in wb.sheetnames:
            wb.close()
            return set()

        ws = wb[SHEET_RIEPILOGO]

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            wb.close()
            return set()

        header_map = {
            str(value).strip(): idx
            for idx, value in enumerate(header_row)
            if value is not None
        }

        id_col_idx = header_map.get("ID")
        stato_col_idx = header_map.get("Stato")

        if id_col_idx is None:
            wb.close()
            return set()

        processed: set[str] = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            invoice_id = row[id_col_idx] if id_col_idx < len(row) else None
            stato = row[stato_col_idx] if stato_col_idx is not None and stato_col_idx < len(row) else None

            if not invoice_id:
                continue
            if str(invoice_id).strip() == "TOTALE":
                continue
            if stato and str(stato).strip().upper() != "OK":
                continue

            processed.add(str(invoice_id).strip())

        wb.close()
        return processed

    except Exception:
        return set()


def is_already_processed(invoice_id: str, processed_ids: set[str]) -> bool:
    """
    True se la fattura è già presente nello storico.
    """
    normalized = (invoice_id or "").strip()
    if not normalized:
        return False
    return normalized in processed_ids
