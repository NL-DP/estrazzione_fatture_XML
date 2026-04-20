from __future__ import annotations

import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from extractor_fatture.config import COLORI, FIELD_MAP_DETTAGLIO, FIELD_MAP_RIEPILOGO

log = logging.getLogger(__name__)

SHEET_RIEPILOGO = "Riepilogo"
SHEET_DETTAGLIO = "Dettaglio"

# Formato data nativo Excel (DD/MM/YYYY visibile, ordinabile come data)
FMT_DATE = "DD/MM/YYYY"
FMT_CURRENCY = '#,##0.00 "EUR"'
FMT_NUMBER = "#,##0.000"
FMT_PERCENT = '0.00"%"'


def _obj_get(obj: Any, key: str, default=None):
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


def _parse_date(value: str):
    """
    Converte una stringa DD/MM/YYYY in oggetto datetime per Excel.
    Restituisce il valore originale se non riconosciuto.
    """
    if not value:
        return value
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt)
        except (ValueError, TypeError):
            pass
    return value


def _ensure_unfrozen(ws) -> None:
    ws.freeze_panes = None


def _result_testata(result: Any):
    return _obj_get(result, "testata", {})


def _result_linee(result: Any):
    linee = _obj_get(result, "linee", [])
    return linee or []


def _init_foglio_riepilogo(ws) -> None:
    ws.row_dimensions[1].height = 28
    h_font = Font(name="Arial", bold=True, color=COLORI["header_fg"], size=11)
    h_fill = PatternFill("solid", fgColor=COLORI["header_bg"])
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, field in enumerate(FIELD_MAP_RIEPILOGO, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field["label"])
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        ws.column_dimensions[get_column_letter(col_idx)].width = field["width"]

    # Autofilter sulla riga header
    ws.auto_filter.ref = ws.dimensions
    _ensure_unfrozen(ws)


def _scrivi_righe_riepilogo(ws, testate: list[Any]) -> None:
    max_r = ws.max_row

    totale_row_idx = None
    for r in range(max_r, 0, -1):
        val = ws.cell(row=r, column=1).value
        if val == "TOTALE":
            totale_row_idx = r
            break

    start_row = totale_row_idx if totale_row_idx else max_r + 1

    for i, testata in enumerate(testate):
        row_idx = start_row + i
        stato = _obj_get(testata, "stato", "")
        is_ok = stato == "OK"
        is_alt = (row_idx % 2 == 0)

        bg = (COLORI["riga_ok_a"] if is_alt else COLORI["riga_ok_b"]) if is_ok else COLORI["riga_err"]
        row_fill = PatternFill("solid", fgColor=bg)
        row_font = Font(
            name="Arial",
            size=10,
            color=COLORI["testo_err"] if not is_ok else "000000",
        )

        for col_idx, field in enumerate(FIELD_MAP_RIEPILOGO, start=1):
            raw_value = _obj_get(testata, field["key"])
            fmt = field.get("format", "text")

            # Converti le date in oggetto datetime per Excel nativo
            if fmt == "date" and isinstance(raw_value, str):
                value = _parse_date(raw_value)
            else:
                value = raw_value

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.font = row_font

            if fmt == "currency" and isinstance(value, (int, float)):
                cell.number_format = FMT_CURRENCY
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif fmt == "date":
                if isinstance(value, datetime):
                    cell.number_format = FMT_DATE
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif fmt == "center":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    nuovo_tot_idx = start_row + len(testate)
    stato_col_idx = next((i for i, f in enumerate(FIELD_MAP_RIEPILOGO, 1) if f["key"] == "stato"), None)
    stato_letter = get_column_letter(stato_col_idx) if stato_col_idx else None

    t_font = Font(name="Arial", bold=True, size=11)
    t_fill = PatternFill("solid", fgColor=COLORI["totale_bg"])

    label_scritto = False
    for col_idx, field in enumerate(FIELD_MAP_RIEPILOGO, start=1):
        cell = ws.cell(row=nuovo_tot_idx, column=col_idx)
        cell.font = t_font
        cell.fill = t_fill

        e_sum = field.get("sumtotal") and field.get("format") == "currency"
        if e_sum and stato_letter:
            col_letter = get_column_letter(col_idx)
            cell.value = f'=SUMIF({stato_letter}:{stato_letter}, "OK", {col_letter}:{col_letter})'
            cell.number_format = FMT_CURRENCY
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif not label_scritto:
            cell.value = "TOTALE"
            cell.alignment = Alignment(horizontal="right", vertical="center")
            label_scritto = True
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Aggiorna autofilter per coprire anche le nuove righe
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELD_MAP_RIEPILOGO))}{nuovo_tot_idx - 1}"


def _init_foglio_dettaglio(ws) -> None:
    for col_idx, field in enumerate(FIELD_MAP_DETTAGLIO, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = field["width"]
    _ensure_unfrozen(ws)


def _scrivi_blocco_fattura(ws, testata: Any, linee: list[Any]) -> None:
    n_cols = len(FIELD_MAP_DETTAGLIO)

    current_row = ws.max_row + 1
    if current_row == 2 and ws.cell(row=1, column=1).value is None:
        current_row = 1

    ws.row_dimensions[current_row].height = 24

    testo_banner = " {} | N. {} | {} | Totale: {:,.2f} {}".format(
        _obj_get(testata, "fornitore", "") or "",
        _obj_get(testata, "numero_fattura", "") or "",
        _obj_get(testata, "data", "") or "",
        _obj_get(testata, "totale", 0.0) or 0.0,
        _obj_get(testata, "divisa", "EUR") or "EUR",
    )

    cell = ws.cell(row=current_row, column=1, value=testo_banner)
    cell.font = Font(name="Arial", bold=True, color=COLORI["sep_fg"], size=11)
    cell.fill = PatternFill("solid", fgColor=COLORI["sep_bg"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=n_cols)
    current_row += 1

    ws.row_dimensions[current_row].height = 18
    h_font = Font(name="Arial", bold=True, color=COLORI["sep_fg"], size=10)
    h_fill = PatternFill("solid", fgColor=COLORI["det_hdr_bg"])

    for col_idx, field in enumerate(FIELD_MAP_DETTAGLIO, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=field["label"])
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row += 1

    for li, linea in enumerate(linee):
        ws.row_dimensions[current_row].height = 15
        bg = COLORI["det_alt_a"] if li % 2 == 0 else COLORI["det_alt_b"]
        row_fill = PatternFill("solid", fgColor=bg)
        row_font = Font(name="Arial", size=10)

        for col_idx, field in enumerate(FIELD_MAP_DETTAGLIO, start=1):
            value = _obj_get(linea, field["key"])
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.fill = row_fill
            cell.font = row_font

            fmt = field.get("format", "text")
            if fmt == "currency" and isinstance(value, (int, float)):
                cell.number_format = FMT_CURRENCY
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif fmt == "number" and isinstance(value, (int, float)):
                cell.number_format = FMT_NUMBER
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif fmt == "percent" and isinstance(value, (int, float)):
                cell.number_format = FMT_PERCENT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif fmt == "center":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        current_row += 1

    ws.row_dimensions[current_row].height = 18
    sub_fill = PatternFill("solid", fgColor=COLORI["det_sub_bg"])
    sub_font = Font(name="Arial", bold=True, size=10)

    cell = ws.cell(row=current_row, column=1, value="Subtotale fattura")
    cell.font = sub_font
    cell.fill = sub_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")

    for col_idx, field in enumerate(FIELD_MAP_DETTAGLIO, start=1):
        if field["key"] == "prezzo_totale":
            cell = ws.cell(row=current_row, column=col_idx, value=_obj_get(testata, "totale"))
            cell.font = sub_font
            cell.fill = sub_fill
            cell.number_format = FMT_CURRENCY
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif field["key"] not in ("codice", "descrizione"):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = sub_fill

    current_row += 1
    ws.row_dimensions[current_row].height = 8


def build_excel(risultati: list[Any], output_path: Path) -> None:
    if output_path.exists():
        wb = load_workbook(str(output_path))
        log.info("Modalita APPEND: carico '%s'", output_path.name)
    else:
        wb = Workbook()
        wb.active.title = SHEET_RIEPILOGO
        log.info("Modalita NUOVO: creo '%s'", output_path.name)

    if SHEET_RIEPILOGO not in wb.sheetnames:
        ws_riepilogo = wb.create_sheet(SHEET_RIEPILOGO, 0)
        _init_foglio_riepilogo(ws_riepilogo)
    else:
        ws_riepilogo = wb[SHEET_RIEPILOGO]
        if ws_riepilogo.max_row == 1 and ws_riepilogo.cell(row=1, column=1).value is None:
            _init_foglio_riepilogo(ws_riepilogo)

    if SHEET_DETTAGLIO not in wb.sheetnames:
        ws_dettaglio = wb.create_sheet(SHEET_DETTAGLIO)
        _init_foglio_dettaglio(ws_dettaglio)
    else:
        ws_dettaglio = wb[SHEET_DETTAGLIO]
        if ws_dettaglio.max_row == 1 and ws_dettaglio.cell(row=1, column=1).value is None:
            _init_foglio_dettaglio(ws_dettaglio)

    _ensure_unfrozen(ws_riepilogo)
    _ensure_unfrozen(ws_dettaglio)

    testate = [_result_testata(r) for r in risultati]
    _scrivi_righe_riepilogo(ws_riepilogo, testate)

    for r in risultati:
        testata = _result_testata(r)
        if _obj_get(testata, "stato") == "OK":
            _scrivi_blocco_fattura(ws_dettaglio, testata, _result_linee(r))

    # Salvataggio sicuro: scrive su file temp poi rinomina
    # Evita crash se Excel ha il file aperto
    tmp_path = output_path.with_suffix(".tmp.xlsx")
    try:
        wb.save(str(tmp_path))
        shutil.move(str(tmp_path), str(output_path))
    except PermissionError:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
        log.error(
            "Impossibile salvare '%s': file aperto in Excel. Chiuderlo e rieseguire.",
            output_path,
        )
        raise

    n_linee = sum(len(_result_linee(r)) for r in risultati if _obj_get(_result_testata(r), "stato") == "OK")
    log.info("Excel salvato: %s (%d fatture, %d righe articolo)", output_path, len(risultati), n_linee)
